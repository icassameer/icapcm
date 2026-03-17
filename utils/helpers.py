# ============================================================
#  ICA PMS — Utility Helpers
# ============================================================

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
import sys, os, datetime

def resource_path(relative_path):
    """Get absolute path for PyInstaller bundled files."""
    if getattr(sys, '_MEIPASS', None):
        base = sys._MEIPASS
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_path)

def format_currency(amount, symbol='₹'):
    try:
        return f"{symbol}{float(amount):,.2f}"
    except:
        return f"{symbol}0.00"

def format_date(dt_str):
    if not dt_str:
        return ''
    try:
        dt = datetime.datetime.strptime(dt_str[:19], '%Y-%m-%d %H:%M:%S')
        return dt.strftime('%d-%b-%Y %H:%M')
    except:
        return dt_str[:10]

def today_str():
    return datetime.date.today().isoformat()

def now_str():
    return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def month_start_str():
    return datetime.date.today().replace(day=1).isoformat()

# ── Reusable Widgets ─────────────────────────────────────────

def make_entry(parent, placeholder='', width=200, show=''):
    e = ctk.CTkEntry(parent, placeholder_text=placeholder, width=width,
                     fg_color='white', border_color='#D1D5DB',
                     text_color='#1A1A2E', show=show)
    return e

def make_label(parent, text, size=13, weight='normal', color='#1A1A2E'):
    return ctk.CTkLabel(parent, text=text, font=('Segoe UI', size, weight), text_color=color)

def make_btn(parent, text, command=None, color='#1B2A6B', hover='#243580', width=120, height=36):
    return ctk.CTkButton(parent, text=text, command=command,
                         fg_color=color, hover_color=hover,
                         width=width, height=height, corner_radius=8,
                         font=('Segoe UI', 13, 'bold'))

def make_green_btn(parent, text, command=None, width=120, height=36):
    return make_btn(parent, text, command, color='#00D084', hover='#00B870',
                    width=width, height=height)

def make_danger_btn(parent, text, command=None, width=120, height=36):
    return make_btn(parent, text, command, color='#EF4444', hover='#DC2626',
                    width=width, height=height)

def card_frame(parent, **kwargs):
    return ctk.CTkFrame(parent, fg_color='white', corner_radius=12,
                        border_width=1, border_color='#E5E7EB', **kwargs)

def stat_card(parent, title, value, subtitle='', color='#1B2A6B', icon=''):
    frame = ctk.CTkFrame(parent, fg_color=color, corner_radius=14, width=180, height=110)
    frame.pack_propagate(False)
    ctk.CTkLabel(frame, text=f"{icon}  {title}" if icon else title,
                 font=('Segoe UI', 11), text_color='#BFCFEE').pack(anchor='w', padx=16, pady=(14,0))
    ctk.CTkLabel(frame, text=str(value),
                 font=('Segoe UI', 26, 'bold'), text_color='white').pack(anchor='w', padx=16, pady=2)
    if subtitle:
        ctk.CTkLabel(frame, text=subtitle,
                     font=('Segoe UI', 10), text_color='#99B0D8').pack(anchor='w', padx=16)
    return frame


class DataTable(tk.Frame):
    """Reusable treeview table — fully responsive, fills available space."""
    def __init__(self, parent, columns, column_widths=None, height=None, **kwargs):
        super().__init__(parent, bg='white', **kwargs)
        self.columns = columns

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('ICA.Treeview',
                         background='white',
                         foreground='#1A1A2E',
                         rowheight=32,
                         fieldbackground='white',
                         font=('Segoe UI', 11))
        style.configure('ICA.Treeview.Heading',
                         background='#1B2A6B',
                         foreground='white',
                         font=('Segoe UI', 11, 'bold'),
                         relief='flat')
        style.map('ICA.Treeview', background=[('selected', '#E0E7FF')])
        style.map('ICA.Treeview.Heading', background=[('active', '#243580')])

        # Scrollbars
        vsb = ttk.Scrollbar(self, orient='vertical')
        hsb = ttk.Scrollbar(self, orient='horizontal')

        # No fixed height — fills whatever space the container gives it
        self.tree = ttk.Treeview(self, columns=columns, show='headings',
                                  style='ICA.Treeview',
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        for col in columns:
            w = (column_widths or {}).get(col, 120)
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=50)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Alternating row colors
        self.tree.tag_configure('odd', background='#F8FAFC')
        self.tree.tag_configure('even', background='white')
        self.tree.tag_configure('low', background='#FEF3C7')
        self.tree.tag_configure('danger', background='#FEE2E2')

    def load(self, rows):
        self.tree.delete(*self.tree.get_children())
        for i, row in enumerate(rows):
            tag = 'odd' if i % 2 else 'even'
            self.tree.insert('', 'end', values=list(row.values())[:len(self.columns)],
                              iid=str(i), tags=(tag,))

    def load_raw(self, rows):
        """Load list of tuples/lists."""
        self.tree.delete(*self.tree.get_children())
        for i, row in enumerate(rows):
            tag = 'odd' if i % 2 else 'even'
            self.tree.insert('', 'end', values=list(row), iid=str(i), tags=(tag,))

    def selected_values(self):
        sel = self.tree.selection()
        if sel:
            return self.tree.item(sel[0])['values']
        return None

    def bind_select(self, callback):
        self.tree.bind('<<TreeviewSelect>>', callback)

    def bind_double(self, callback):
        self.tree.bind('<Double-1>', callback)


class SearchBar(ctk.CTkFrame):
    def __init__(self, parent, placeholder='Search...', on_change=None, width=300):
        super().__init__(parent, fg_color='transparent')
        self._callback = on_change
        self._var = ctk.StringVar()
        self._var.trace_add('write', self._on_change)

        ctk.CTkLabel(self, text='🔍', font=('Segoe UI', 14)).pack(side='left', padx=(0,4))
        self.entry = ctk.CTkEntry(self, textvariable=self._var,
                                   placeholder_text=placeholder, width=width,
                                   fg_color='white', border_color='#D1D5DB',
                                   text_color='#1A1A2E')
        self.entry.pack(side='left')

    def _on_change(self, *_):
        if self._callback:
            self._callback(self._var.get())

    def get(self):
        return self._var.get()


def show_toast(parent, message, color='#10B981', duration=2500):
    toast = ctk.CTkToplevel(parent)
    toast.overrideredirect(True)
    toast.attributes('-topmost', True)
    px = parent.winfo_rootx() + parent.winfo_width() // 2 - 150
    py = parent.winfo_rooty() + parent.winfo_height() - 80
    toast.geometry(f"300x50+{px}+{py}")
    ctk.CTkLabel(toast, text=message, fg_color=color, text_color='white',
                  font=('Segoe UI', 12, 'bold'), corner_radius=8).pack(fill='both', expand=True)
    toast.after(duration, toast.destroy)


def confirm_dialog(parent, title, message):
    dialog = ctk.CTkToplevel(parent)
    dialog.title(title)
    dialog.geometry("400x160")
    dialog.resizable(False, False)
    dialog.grab_set()

    result = [False]

    ctk.CTkLabel(dialog, text=message, font=('Segoe UI', 13),
                  wraplength=360).pack(pady=(24, 16))

    btn_frame = ctk.CTkFrame(dialog, fg_color='transparent')
    btn_frame.pack()

    def on_yes():
        result[0] = True
        dialog.destroy()

    make_danger_btn(btn_frame, 'Yes, Delete', on_yes, width=130).pack(side='left', padx=8)
    make_btn(btn_frame, 'Cancel', dialog.destroy, width=110).pack(side='left', padx=8)

    dialog.wait_window()
    return result[0]


# ── Export Helpers ───────────────────────────────────────────

def export_to_excel(data: list, columns: list, filepath: str, sheet_name='Report'):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row
        header_fill = PatternFill("solid", fgColor="1B2A6B")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_idx, row in enumerate(data, 2):
            if isinstance(row, dict):
                values = [row.get(c, '') for c in columns]
            else:
                values = list(row)
            for col_idx, val in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            # Alternate row color
            if row_idx % 2 == 0:
                fill = PatternFill("solid", fgColor="F0F2F8")
                for col_idx in range(1, len(columns)+1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Excel export error: {e}")
        return False


def export_to_pdf(title, headers, rows, filepath, company_name='ICA – Integrating Consulting & Automation'):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm

        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                                 leftMargin=1*cm, rightMargin=1*cm,
                                 topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle('title', parent=styles['Title'],
                                      fontSize=16, textColor=colors.HexColor('#1B2A6B'))
        elements.append(Paragraph(company_name, title_style))
        elements.append(Paragraph(title, styles['Heading2']))
        elements.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}",
                                   styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        # Table
        table_data = [headers] + [list(r) for r in rows]
        col_count = len(headers)
        available_width = landscape(A4)[0] - 2*cm
        col_width = available_width / col_count

        t = Table(table_data, colWidths=[col_width]*col_count, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B2A6B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F0F2F8')]),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWHEIGHT', (0,0), (-1,-1), 20),
        ]))
        elements.append(t)
        doc.build(elements)
        return True
    except Exception as e:
        print(f"PDF export error: {e}")
        return False
